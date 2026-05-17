#!/usr/bin/env python3
"""
SimpleScalar sim-outorder Results Table Generator
===================================================
COA/CA Project (Spring 2026) — ITU Lahore

Supports Part I, II, and III result files.

Usage:
    python simscalar_tables.py

The script will:
  - Ask the user for benchmark name(s) and result file paths
  - Parse all sim-outorder result files
  - Generate per-benchmark tables, a cross-benchmark comparison table
  - Generate additional summary tables: Final Results, Key Outputs, Important Findings
  - Save everything to a nicely formatted Excel (.xlsx) file and print to console
"""

import re
import os
import sys

# ─────────────────────────────────────────────────────────────────────────────
#  PARSER  (same patterns as simscalar_grapher.py for consistency)
# ─────────────────────────────────────────────────────────────────────────────

STAT_PATTERNS = {
    # Core performance
    "sim_num_insn":     r"sim_num_insn\s+([\d.e+]+)",
    "sim_num_refs":     r"sim_num_refs\s+([\d.e+]+)",
    "sim_num_loads":    r"sim_num_loads\s+([\d.e+]+)",
    "sim_num_stores":   r"sim_num_stores\s+([\d.e+]+)",
    "sim_num_branches": r"sim_num_branches\s+([\d.e+]+)",
    "sim_cycle":        r"sim_cycle\s+([\d.e+]+)",
    "sim_IPC":          r"sim_IPC\s+([\d.e+]+)",
    "sim_CPI":          r"sim_CPI\s+([\d.e+]+)",
    "sim_inst_rate":    r"sim_inst_rate\s+([\d.e+]+)",
    "sim_exec_BW":      r"sim_exec_BW\s+([\d.e+]+)",
    "sim_IPB":          r"sim_IPB\s+([\d.e+]+)",
    "avg_sim_slip":     r"avg_sim_slip\s+([\d.e+]+)",

    # IFQ
    "ifq_occupancy":    r"ifq_occupancy\s+([\d.e+]+)",
    "ifq_rate":         r"ifq_rate\s+([\d.e+]+)",
    "ifq_latency":      r"ifq_latency\s+([\d.e+]+)",
    "ifq_full":         r"ifq_full\s+([\d.e+]+)",

    # RUU / LSQ
    "ruu_occupancy":    r"ruu_occupancy\s+([\d.e+]+)",
    "ruu_full":         r"ruu_full\s+([\d.e+]+)",
    "lsq_occupancy":    r"lsq_occupancy\s+([\d.e+]+)",
    "lsq_full":         r"lsq_full\s+([\d.e+]+)",

    # Branch predictor
    "bpred_lookups":    r"bpred[\w.]+\.lookups\s+([\d.e+]+)",
    "bpred_updates":    r"bpred[\w.]+\.updates\s+([\d.e+]+)",
    "bpred_addr_hits":  r"bpred[\w.]+\.addr_hits\s+([\d.e+]+)",
    "bpred_dir_hits":   r"bpred[\w.]+\.dir_hits\s+([\d.e+]+)",
    "bpred_misses":     r"bpred[\w.]+\.misses\s+([\d.e+]+)",
    "bpred_addr_rate":  r"bpred[\w.]+\.bpred_addr_rate\s+([\d.e+]+)",
    "bpred_dir_rate":   r"bpred[\w.]+\.bpred_dir_rate\s+([\d.e+]+)",
    "bpred_jr_rate":    r"bpred[\w.]+\.bpred_jr_rate\s+([\d.e+]+)",

    # DL1 / IL1 / DL2 / IL2
    "dl1_accesses":     r"dl1\.accesses\s+([\d.e+]+)",
    "dl1_hits":         r"dl1\.hits\s+([\d.e+]+)",
    "dl1_misses":       r"dl1\.misses\s+([\d.e+]+)",
    "dl1_miss_rate":    r"dl1\.miss_rate\s+([\d.e+]+)",
    "dl1_repl_rate":    r"dl1\.repl_rate\s+([\d.e+]+)",
    "dl1_wb_rate":      r"dl1\.wb_rate\s+([\d.e+]+)",

    "il1_accesses":     r"il1\.accesses\s+([\d.e+]+)",
    "il1_hits":         r"il1\.hits\s+([\d.e+]+)",
    "il1_misses":       r"il1\.misses\s+([\d.e+]+)",
    "il1_miss_rate":    r"il1\.miss_rate\s+([\d.e+]+)",

    "dl2_accesses":     r"dl2\.accesses\s+([\d.e+]+)",
    "dl2_hits":         r"dl2\.hits\s+([\d.e+]+)",
    "dl2_misses":       r"dl2\.misses\s+([\d.e+]+)",
    "dl2_miss_rate":    r"dl2\.miss_rate\s+([\d.e+]+)",

    "il2_accesses":     r"il2\.accesses\s+([\d.e+]+)",
    "il2_hits":         r"il2\.hits\s+([\d.e+]+)",
    "il2_misses":       r"il2\.misses\s+([\d.e+]+)",
    "il2_miss_rate":    r"il2\.miss_rate\s+([\d.e+]+)",

    # TLB
    "itlb_miss_rate":   r"itlb\.miss_rate\s+([\d.e+]+)",
    "dtlb_miss_rate":   r"dtlb\.miss_rate\s+([\d.e+]+)",

    # Memory
    "mem_page_count":   r"mem\.page_count\s+([\d.e+]+)",

    # Config
    "fetch_ifqsize":    r"-fetch:ifqsize\s+(\d+)",
    "decode_width":     r"-decode:width\s+(\d+)",
    "issue_width":      r"-issue:width\s+(\d+)",
    "commit_width":     r"-commit:width\s+(\d+)",
    "ruu_size":         r"-ruu:size\s+(\d+)",
    "lsq_size":         r"-lsq:size\s+(\d+)",
}

# Human-readable display names for each stat
STAT_LABELS = {
    "sim_num_insn":     "Total Instructions",
    "sim_num_refs":     "Memory References",
    "sim_num_loads":    "Load Instructions",
    "sim_num_stores":   "Store Instructions",
    "sim_num_branches": "Branch Instructions",
    "sim_cycle":        "Total Cycles",
    "sim_IPC":          "IPC (Instr/Cycle)",
    "sim_CPI":          "CPI (Cycles/Instr)",
    "sim_inst_rate":    "Sim Speed (insts/sec)",
    "sim_exec_BW":      "Exec Bandwidth",
    "sim_IPB":          "Instructions/Branch",
    "avg_sim_slip":     "Avg Sim Slip",
    "ifq_occupancy":    "IFQ Avg Occupancy",
    "ifq_rate":         "IFQ Rate",
    "ifq_latency":      "IFQ Latency",
    "ifq_full":         "IFQ Full (fraction)",
    "ruu_occupancy":    "RUU Avg Occupancy",
    "ruu_full":         "RUU Full (fraction)",
    "lsq_occupancy":    "LSQ Avg Occupancy",
    "lsq_full":         "LSQ Full (fraction)",
    "bpred_lookups":    "Branch Pred Lookups",
    "bpred_updates":    "Branch Pred Updates",
    "bpred_addr_hits":  "Branch Addr Hits (BTB)",
    "bpred_dir_hits":   "Branch Dir Hits",
    "bpred_misses":     "Branch Pred Misses",
    "bpred_addr_rate":  "Branch Addr Rate",
    "bpred_dir_rate":   "Branch Dir Rate",
    "bpred_jr_rate":    "Branch JR Rate",
    "dl1_accesses":     "DL1 Accesses",
    "dl1_hits":         "DL1 Hits",
    "dl1_misses":       "DL1 Misses",
    "dl1_miss_rate":    "DL1 Miss Rate",
    "dl1_repl_rate":    "DL1 Replace Rate",
    "dl1_wb_rate":      "DL1 Writeback Rate",
    "il1_accesses":     "IL1 Accesses",
    "il1_hits":         "IL1 Hits",
    "il1_misses":       "IL1 Misses",
    "il1_miss_rate":    "IL1 Miss Rate",
    "dl2_accesses":     "DL2 Accesses",
    "dl2_hits":         "DL2 Hits",
    "dl2_misses":       "DL2 Misses",
    "dl2_miss_rate":    "DL2 Miss Rate",
    "il2_accesses":     "IL2 Accesses",
    "il2_hits":         "IL2 Hits",
    "il2_misses":       "IL2 Misses",
    "il2_miss_rate":    "IL2 Miss Rate",
    "itlb_miss_rate":   "ITLB Miss Rate",
    "dtlb_miss_rate":   "DTLB Miss Rate",
    "mem_page_count":   "Memory Pages Used",
    "fetch_ifqsize":    "IFQ Size (config)",
    "decode_width":     "Decode Width (config)",
    "issue_width":      "Issue Width (config)",
    "commit_width":     "Commit Width (config)",
    "ruu_size":         "RUU Size (config)",
    "lsq_size":         "LSQ Size (config)",
}

# Groups for organised display
TABLE_GROUPS = {
    "Core Performance": [
        "sim_num_insn", "sim_num_branches", "sim_cycle",
        "sim_IPC", "sim_CPI", "sim_inst_rate",
        "sim_num_loads", "sim_num_stores", "sim_num_refs",
    ],
    "Cache — DL1": [
        "dl1_accesses", "dl1_hits", "dl1_misses", "dl1_miss_rate",
        "dl1_repl_rate", "dl1_wb_rate",
    ],
    "Cache — IL1": [
        "il1_accesses", "il1_hits", "il1_misses", "il1_miss_rate",
    ],
    "Cache — DL2": [
        "dl2_accesses", "dl2_hits", "dl2_misses", "dl2_miss_rate",
    ],
    "Cache — IL2": [
        "il2_accesses", "il2_hits", "il2_misses", "il2_miss_rate",
    ],
    "Branch Predictor": [
        "bpred_lookups", "bpred_updates", "bpred_dir_hits",
        "bpred_addr_hits", "bpred_misses",
        "bpred_dir_rate", "bpred_addr_rate", "bpred_jr_rate",
    ],
    "Pipeline Buffers": [
        "ifq_occupancy", "ifq_rate", "ifq_latency", "ifq_full",
        "ruu_occupancy", "ruu_full",
        "lsq_occupancy", "lsq_full",
        "avg_sim_slip",
    ],
    "TLB & Memory": [
        "itlb_miss_rate", "dtlb_miss_rate", "mem_page_count",
    ],
    "Configuration": [
        "fetch_ifqsize", "decode_width", "issue_width",
        "commit_width", "ruu_size", "lsq_size",
    ],
}

# Key stats shown in summary/comparison tables
KEY_STATS = [
    "sim_cycle", "sim_IPC", "sim_CPI",
    "dl1_miss_rate", "il1_miss_rate", "dl2_miss_rate",
    "bpred_dir_rate", "bpred_misses",
    "ifq_occupancy", "ruu_occupancy", "lsq_occupancy",
]

# ─────────────────────────────────────────────────────────────────────────────
#  PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_file(filepath):
    with open(filepath, "r", errors="replace") as f:
        text = f.read()
    stats = {}
    for key, pattern in STAT_PATTERNS.items():
        m = re.search(pattern, text, re.MULTILINE)
        if m:
            stats[key] = float(m.group(1))
    # Derived hit rates
    for base in ("dl1", "il1", "dl2", "il2"):
        mr_key = f"{base}_miss_rate"
        hr_key = f"{base}_hit_rate"
        if mr_key in stats:
            stats[hr_key] = 1.0 - stats[mr_key]
    if "bpred_misses" in stats and "bpred_updates" in stats and stats["bpred_updates"] > 0:
        stats["bpred_miss_rate"] = stats["bpred_misses"] / stats["bpred_updates"]
    return stats


# ─────────────────────────────────────────────────────────────────────────────
#  FORMAT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def fmt(v, key=""):
    if v is None:
        return "N/A"
    if isinstance(v, float):
        if "rate" in key or "miss_rate" in key or "hit_rate" in key or key in (
                "sim_IPC", "sim_CPI", "ifq_occupancy", "ruu_occupancy",
                "lsq_occupancy", "avg_sim_slip", "ifq_rate", "ifq_latency",
                "ifq_full", "ruu_full", "lsq_full"):
            if v < 0.001:
                return f"{v:.6f}"
            return f"{v:.4f}"
        if v >= 1e9:
            return f"{v/1e9:.3f} B"
        if v >= 1e6:
            return f"{v/1e6:.3f} M"
        if v >= 1e3:
            return f"{v:,.0f}"
        return f"{v:.4f}"
    return str(int(v))


# ─────────────────────────────────────────────────────────────────────────────
#  CONSOLE TABLE PRINTER
# ─────────────────────────────────────────────────────────────────────────────

def print_separator(widths, char="─", junc="┼"):
    return "├" + junc.join(char * w for w in widths) + "┤"

def print_header_sep(widths):
    return "╞" + "╪".join("═" * w for w in widths) + "╡"

def print_table(title, col_headers, rows, indent=0):
    """
    rows: list of (label, [val1, val2, ...]) tuples
    """
    pad = " " * indent
    ncols = len(col_headers)

    # Compute column widths
    col_w = [len(h) + 2 for h in col_headers]
    label_w = max((len(r[0]) for r in rows), default=20) + 2
    for _, vals in rows:
        for i, v in enumerate(vals):
            col_w[i] = max(col_w[i], len(str(v)) + 2)

    total_w = label_w + sum(col_w) + ncols + 1

    bar = "═" * total_w
    print(f"\n{pad}╔{bar}╗")
    title_str = f" {title} "
    print(f"{pad}║{title_str:^{total_w}}║")
    print(f"{pad}╠{bar}╣")

    # Header row
    header_row = f"{'Metric':<{label_w}}" + "│".join(f"{h:^{w}}" for h, w in zip(col_headers, col_w))
    print(f"{pad}║ {header_row}║")
    print(f"{pad}╠{'═'*label_w}{'╪' + '╪'.join('═'*w for w in col_w)}╣")

    prev_group = None
    for label, vals in rows:
        # Group separator
        if "──" in label:
            print(f"{pad}╟{'─'*(label_w)}{'┼' + '┼'.join('─'*w for w in col_w)}╢")
            group_str = f" {label.strip('─ ')} "
            print(f"{pad}║{group_str:<{label_w + sum(col_w) + ncols}}║")
            print(f"{pad}╟{'─'*(label_w)}{'┼' + '┼'.join('─'*w for w in col_w)}╢")
            continue
        row_str = f"{label:<{label_w}}" + "│".join(f"{str(v):>{w-1}} " for v, w in zip(vals, col_w))
        print(f"{pad}║ {row_str}║")

    print(f"{pad}╚{'═'*total_w}╝")


# ─────────────────────────────────────────────────────────────────────────────
#  BUILD TABLE ROWS
# ─────────────────────────────────────────────────────────────────────────────

def build_full_table_rows(all_stats_list):
    """Returns rows [(label, [val, ...]), ...] grouped by category."""
    rows = []
    for group_name, stat_keys in TABLE_GROUPS.items():
        rows.append((f"── {group_name} ", [""] * len(all_stats_list)))
        for key in stat_keys:
            vals_present = [s.get(key) for s in all_stats_list]
            if all(v is None for v in vals_present):
                continue
            label = STAT_LABELS.get(key, key)
            vals = [fmt(v, key) for v in vals_present]
            rows.append((label, vals))
    return rows


def build_comparison_rows(benchmark_stats_map):
    """
    benchmark_stats_map: {benchmark_name: {config_label: stats_dict}}
    Returns rows for cross-benchmark comparison on key stats only.
    """
    rows = []
    benchmarks = list(benchmark_stats_map.keys())
    for key in KEY_STATS:
        label = STAT_LABELS.get(key, key)
        bench_vals = []
        for bname in benchmarks:
            configs = benchmark_stats_map[bname]
            # Use the first config's value for a representative number
            # (all configs for cross-benchmark comparison side by side)
            vals = [configs[c].get(key) for c in configs]
            if all(v is None for v in vals):
                bench_vals.append("N/A")
            else:
                # Show min → max if multiple configs
                non_none = [v for v in vals if v is not None]
                if len(non_none) == 1:
                    bench_vals.append(fmt(non_none[0], key))
                else:
                    bench_vals.append(f"{fmt(min(non_none), key)} → {fmt(max(non_none), key)}")
        rows.append((label, bench_vals))
    return rows, benchmarks


def derive_findings(benchmark_name, config_labels, all_stats_list):
    """Compute derived analysis values for the findings table."""
    findings = []

    # IPC range
    ipcs = [s.get("sim_IPC") for s in all_stats_list if s.get("sim_IPC") is not None]
    if len(ipcs) >= 2:
        best_label = config_labels[all_stats_list.index(max(all_stats_list, key=lambda s: s.get("sim_IPC", 0)))]
        worst_label = config_labels[all_stats_list.index(min(all_stats_list, key=lambda s: s.get("sim_IPC", 0)))]
        findings.append(("Best IPC Config",      [best_label]))
        findings.append(("Worst IPC Config",     [worst_label]))
        findings.append(("IPC Range",            [f"{min(ipcs):.4f} – {max(ipcs):.4f}"]))
        speedup = max(ipcs) / min(ipcs) if min(ipcs) > 0 else 0
        findings.append(("IPC Speedup (best/worst)", [f"{speedup:.2f}×"]))

    # Miss rates
    for cache in ("dl1", "il1", "dl2"):
        key = f"{cache}_miss_rate"
        mrs = [s.get(key) for s in all_stats_list if s.get(key) is not None]
        if mrs:
            findings.append((f"{cache.upper()} Miss Rate Range",
                             [f"{min(mrs)*100:.3f}% – {max(mrs)*100:.3f}%"]))

    # Branch
    bdr = [s.get("bpred_dir_rate") for s in all_stats_list if s.get("bpred_dir_rate") is not None]
    if bdr:
        findings.append(("Branch Dir Rate Range",
                         [f"{min(bdr)*100:.3f}% – {max(bdr)*100:.3f}%"]))

    # Cycle overhead across configs
    cycles = [s.get("sim_cycle") for s in all_stats_list if s.get("sim_cycle") is not None]
    if len(cycles) >= 2:
        overhead = (max(cycles) - min(cycles)) / min(cycles) * 100
        findings.append(("Cycle Overhead (max vs min)", [f"{overhead:.1f}%"]))

    return findings


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def write_excel(output_path, benchmark_data, comparison_rows, benchmark_names,
                summary_tables):
    """
    benchmark_data: [(benchmark_name, config_labels, all_stats_list), ...]
    comparison_rows: [(label, [val_per_benchmark, ...]), ...]
    benchmark_names: list of str
    summary_tables: [(sheet_title, col_headers, rows), ...]
    """
    try:
        import openpyxl
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n  ⚠  openpyxl not installed — skipping Excel output.")
        print("     Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Color palette ──
    DARK_BLUE  = "1F3864"   # header bg
    MID_BLUE   = "2E75B6"   # subheader / group
    LIGHT_BLUE = "D6E4F0"   # group label row
    ALT_ROW    = "EBF3FB"   # alternating row
    WHITE      = "FFFFFF"
    GOLD       = "F5C842"   # highlight best value
    RED_LIGHT  = "FCE4D6"   # highlight worst
    TEXT_WHITE = "FFFFFF"
    TEXT_DARK  = "1A1A2E"

    thin = Side(style="thin", color="AAAAAA")
    med  = Side(style="medium", color=MID_BLUE)
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_med_bottom = Border(left=thin, right=thin, top=thin,
                                bottom=Side(style="medium", color=MID_BLUE))

    def hdr_fill(color=DARK_BLUE):
        return PatternFill("solid", fgColor=color)

    def cell_fill(color):
        return PatternFill("solid", fgColor=color)

    def set_col_widths(ws, widths_dict):
        for col_letter, w in widths_dict.items():
            ws.column_dimensions[col_letter].width = w

    def style_header_row(ws, row_num, num_cols, bg=DARK_BLUE, fg=TEXT_WHITE,
                          bold=True, size=11, center=True):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = hdr_fill(bg)
            cell.font = Font(color=fg, bold=bold, name="Calibri", size=size)
            cell.border = border_thin
            if center:
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)

    def style_data_row(ws, row_num, num_cols, alt=False, is_group=False):
        bg = LIGHT_BLUE if is_group else (ALT_ROW if alt else WHITE)
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = cell_fill(bg)
            cell.border = border_thin
            cell.font = Font(name="Calibri", size=10,
                             bold=is_group, color=MID_BLUE if is_group else TEXT_DARK)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if c > 1 else "left",
                                       wrap_text=False)

    # ═════════════════════════════════════════════════════════════════════════
    #  Per-benchmark sheets
    # ═════════════════════════════════════════════════════════════════════════
    for (bname, config_labels, all_stats_list) in benchmark_data:
        ws = wb.create_sheet(title=bname[:28])   # sheet name limit

        # ── Title banner ──
        num_cols = 1 + len(config_labels)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=num_cols)
        ws.row_dimensions[1].height = 28
        title_cell = ws.cell(row=1, column=1,
                             value=f"Benchmark: {bname}  │  sim-outorder Results")
        title_cell.fill = hdr_fill(DARK_BLUE)
        title_cell.font = Font(color=TEXT_WHITE, bold=True, size=13, name="Calibri")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Sub-header (column names) ──
        ws.cell(row=2, column=1, value="Metric")
        for ci, lbl in enumerate(config_labels, start=2):
            ws.cell(row=2, column=ci, value=lbl)
        ws.row_dimensions[2].height = 22
        style_header_row(ws, 2, num_cols, bg=MID_BLUE)

        # ── Data rows ──
        table_rows = build_full_table_rows(all_stats_list)
        row_num = 3
        alt_flag = False
        for label, vals in table_rows:
            is_group = label.startswith("──")
            ws.row_dimensions[row_num].height = 18 if not is_group else 20

            if is_group:
                ws.merge_cells(start_row=row_num, start_column=1,
                               end_row=row_num, end_column=num_cols)
                ws.cell(row=row_num, column=1,
                        value=f"  {label.strip('─ ')}")
                style_data_row(ws, row_num, num_cols, is_group=True)
                alt_flag = False
            else:
                ws.cell(row=row_num, column=1, value=f"  {label}")
                for ci, v in enumerate(vals, start=2):
                    ws.cell(row=row_num, column=ci, value=v)
                style_data_row(ws, row_num, num_cols, alt=alt_flag)
                alt_flag = not alt_flag

            row_num += 1

        # ── Column widths ──
        col_widths = {"A": 32}
        for ci in range(2, num_cols + 1):
            col_widths[get_column_letter(ci)] = max(14, len(config_labels[ci-2]) + 4)
        set_col_widths(ws, col_widths)
        ws.freeze_panes = "B3"

    # ═════════════════════════════════════════════════════════════════════════
    #  Cross-benchmark comparison sheet
    # ═════════════════════════════════════════════════════════════════════════
    if len(benchmark_names) > 1:
        ws = wb.create_sheet(title="Comparison")
        num_cols = 1 + len(benchmark_names)

        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=num_cols)
        ws.row_dimensions[1].height = 28
        c = ws.cell(row=1, column=1, value="Cross-Benchmark Comparison — Key Metrics")
        c.fill = hdr_fill(DARK_BLUE)
        c.font = Font(color=TEXT_WHITE, bold=True, size=13, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=2, column=1, value="Metric")
        for ci, bname in enumerate(benchmark_names, start=2):
            ws.cell(row=2, column=ci, value=bname)
        style_header_row(ws, 2, num_cols, bg=MID_BLUE)
        ws.row_dimensions[2].height = 22

        for ri, (label, vals) in enumerate(comparison_rows, start=3):
            ws.row_dimensions[ri].height = 18
            ws.cell(row=ri, column=1, value=f"  {label}")
            for ci, v in enumerate(vals, start=2):
                ws.cell(row=ri, column=ci, value=v)
            style_data_row(ws, ri, num_cols, alt=(ri % 2 == 0))

        col_widths = {"A": 32}
        for ci in range(2, num_cols + 1):
            col_widths[get_column_letter(ci)] = max(16, len(benchmark_names[ci-2]) + 6)
        set_col_widths(ws, col_widths)
        ws.freeze_panes = "B3"

    # ═════════════════════════════════════════════════════════════════════════
    #  Summary tables (Final Results / Key Outputs / Important Findings)
    # ═════════════════════════════════════════════════════════════════════════
    SUMMARY_COLORS = [("2E75B6", "D6E4F0"), ("1F5C2E", "D9EAD3"), ("6C3483", "E8DAEF")]
    for si, (sheet_title, col_headers, s_rows) in enumerate(summary_tables):
        hdr_col, alt_col = SUMMARY_COLORS[si % len(SUMMARY_COLORS)]
        ws = wb.create_sheet(title=sheet_title[:28])
        num_cols = 1 + len(col_headers)

        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=num_cols)
        ws.row_dimensions[1].height = 28
        c = ws.cell(row=1, column=1, value=sheet_title)
        c.fill = hdr_fill(DARK_BLUE)
        c.font = Font(color=TEXT_WHITE, bold=True, size=13, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=2, column=1, value="Metric / Item")
        for ci, hdr in enumerate(col_headers, start=2):
            ws.cell(row=2, column=ci, value=hdr)
        style_header_row(ws, 2, num_cols, bg=hdr_col)
        ws.row_dimensions[2].height = 22

        for ri, (label, vals) in enumerate(s_rows, start=3):
            ws.row_dimensions[ri].height = 18
            ws.cell(row=ri, column=1, value=f"  {label}")
            for ci, v in enumerate(vals, start=2):
                ws.cell(row=ri, column=ci, value=str(v))
            # alternating color
            bg = alt_col if ri % 2 == 0 else WHITE
            for c_idx in range(1, num_cols + 1):
                cell = ws.cell(row=ri, column=c_idx)
                cell.fill = cell_fill(bg)
                cell.border = border_thin
                cell.font = Font(name="Calibri", size=10, color=TEXT_DARK)
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if c_idx > 1 else "left")

        ws.column_dimensions["A"].width = 38
        for ci in range(2, num_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(
                18, len(col_headers[ci-2]) + 4)
        ws.freeze_panes = "B3"

    wb.save(output_path)
    print(f"\n  ✓  Excel report saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
#  INTERACTIVE INPUT
# ─────────────────────────────────────────────────────────────────────────────

CONFIG_LABELS_DEFAULT = ["Small", "Medium", "Large", "Excessive"]

def prompt_benchmark():
    """Ask for one benchmark's files. Returns (bname, config_labels, all_stats)."""
    print("\n" + "─"*60)
    bname = input("  Benchmark name (e.g. ijpeg, vortex, gcc): ").strip()
    if not bname:
        bname = "Benchmark"

    print(f"\n  Enter result files for '{bname}'.")
    print("  Press ENTER on empty file path when done (min 1 file).\n")

    configs = []
    while True:
        idx = len(configs)
        default_label = CONFIG_LABELS_DEFAULT[idx] if idx < len(CONFIG_LABELS_DEFAULT) else f"Config{idx+1}"
        path = input(f"  File {idx+1} path (or ENTER to finish): ").strip()
        if not path:
            if not configs:
                print("  Please enter at least one file.")
                continue
            break
        if not os.path.isfile(path):
            print(f"  ✗ File not found: {path}")
            continue
        label = input(f"  Label for this file [{default_label}]: ").strip()
        if not label:
            label = default_label
        stats = parse_file(path)
        found = sum(1 for v in stats.values() if v is not None)
        print(f"  ✓  Parsed {found} stats from '{os.path.basename(path)}'")
        configs.append((label, stats))

    labels     = [c[0] for c in configs]
    stats_list = [c[1] for c in configs]
    return bname, labels, stats_list


# ─────────────────────────────────────────────────────────────────────────────
#  SUMMARY TABLE BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def build_final_results_table(benchmark_data):
    """Sheet 1 of extras: Final Results — best IPC, cycles, miss rates per benchmark."""
    col_headers = [bname for bname, _, _ in benchmark_data]
    metrics = [
        ("sim_IPC", "Best IPC"),
        ("sim_CPI", "Best CPI (lowest)"),
        ("sim_cycle", "Min Total Cycles"),
        ("dl1_miss_rate", "Min DL1 Miss Rate"),
        ("il1_miss_rate", "Min IL1 Miss Rate"),
        ("bpred_dir_rate", "Best Branch Dir Rate"),
        ("bpred_misses",  "Min Branch Misses"),
        ("ifq_occupancy", "Avg IFQ Occupancy"),
        ("ruu_occupancy", "Avg RUU Occupancy"),
        ("lsq_occupancy", "Avg LSQ Occupancy"),
    ]
    rows = []
    agg_fn = {
        "sim_IPC": max, "sim_CPI": min, "sim_cycle": min,
        "dl1_miss_rate": min, "il1_miss_rate": min,
        "bpred_dir_rate": max, "bpred_misses": min,
        "ifq_occupancy": lambda x: sum(x)/len(x),
        "ruu_occupancy": lambda x: sum(x)/len(x),
        "lsq_occupancy": lambda x: sum(x)/len(x),
    }
    for stat_key, label in metrics:
        vals = []
        for _, _, stats_list in benchmark_data:
            candidates = [s.get(stat_key) for s in stats_list if s.get(stat_key) is not None]
            if candidates:
                fn = agg_fn.get(stat_key, max)
                vals.append(fmt(fn(candidates), stat_key))
            else:
                vals.append("N/A")
        rows.append((label, vals))
    return "Final Results", col_headers, rows


def build_key_outputs_table(benchmark_data):
    """Sheet 2 of extras: Key Outputs per config per benchmark."""
    # Flatten: rows = one per config across all benchmarks
    rows = []
    col_headers = ["Benchmark", "Config", "IPC", "CPI", "Cycles",
                   "DL1 Miss%", "BP Dir Rate", "RUU Occ"]
    for bname, config_labels, stats_list in benchmark_data:
        for lbl, s in zip(config_labels, stats_list):
            row_vals = [
                bname,
                lbl,
                fmt(s.get("sim_IPC"), "sim_IPC"),
                fmt(s.get("sim_CPI"), "sim_CPI"),
                fmt(s.get("sim_cycle"), "sim_cycle"),
                fmt(s.get("dl1_miss_rate"), "dl1_miss_rate"),
                fmt(s.get("bpred_dir_rate"), "bpred_dir_rate"),
                fmt(s.get("ruu_occupancy"), "ruu_occupancy"),
            ]
            rows.append((f"{bname} / {lbl}", row_vals))
    # Reformat for generic table: label | values
    # col_headers for the sheet = the data columns (after label)
    col_hdrs = ["Benchmark", "Config", "IPC", "CPI", "Cycles",
                "DL1 Miss%", "BP Dir Rate", "RUU Occ"]
    reformat = [(r[0], r[1]) for r in rows]
    return "Key Outputs", col_hdrs, reformat


def build_important_findings_table(benchmark_data):
    """Sheet 3 of extras: Important values / findings per benchmark."""
    col_headers = [bname for bname, _, _ in benchmark_data]
    finding_rows = []

    all_findings_keys = [
        "Best IPC Config",
        "Worst IPC Config",
        "IPC Range",
        "IPC Speedup (best/worst)",
        "DL1 Miss Rate Range",
        "IL1 Miss Rate Range",
        "DL2 Miss Rate Range",
        "Branch Dir Rate Range",
        "Cycle Overhead (max vs min)",
    ]

    per_benchmark = {}
    for bname, config_labels, stats_list in benchmark_data:
        f = dict(derive_findings(bname, config_labels, stats_list))
        per_benchmark[bname] = f

    for key in all_findings_keys:
        vals = []
        for bname, _, _ in benchmark_data:
            f = per_benchmark.get(bname, {})
            v = f.get(key)
            vals.append(v[0] if v else "N/A")
        finding_rows.append((key, vals))

    return "Important Findings", col_headers, finding_rows


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "="*60)
    print("  SimpleScalar Results Table Generator")
    print("  COA/CA Project — Spring 2026 | ITU Lahore")
    print("="*60)

    benchmark_data = []   # list of (bname, config_labels, stats_list)

    print("\nHow many benchmarks do you want to process?")
    while True:
        nb = input("  Number of benchmarks [1]: ").strip()
        nb = nb if nb else "1"
        try:
            nb = int(nb)
            if nb < 1:
                raise ValueError
            break
        except ValueError:
            print("  Please enter a positive integer.")

    for i in range(nb):
        print(f"\n{'═'*60}")
        print(f"  BENCHMARK {i+1} of {nb}")
        bname, labels, stats_list = prompt_benchmark()
        benchmark_data.append((bname, labels, stats_list))

    # ── Per-benchmark console tables ──
    for bname, config_labels, stats_list in benchmark_data:
        rows = build_full_table_rows(stats_list)
        print_table(f"Benchmark: {bname}  |  Full Results",
                    config_labels, rows)

    # ── Cross-benchmark comparison ──
    bmap = {}
    for bname, config_labels, stats_list in benchmark_data:
        bmap[bname] = dict(zip(config_labels, stats_list))
    comp_rows, bnames = build_comparison_rows(bmap)

    if len(benchmark_data) > 1:
        print_table("Cross-Benchmark Comparison — Key Metrics",
                    bnames, comp_rows)
    else:
        comp_rows  = []
        bnames     = []

    # ── Extra summary tables ──
    fr_title, fr_hdrs, fr_rows = build_final_results_table(benchmark_data)
    ko_title, ko_hdrs, ko_rows = build_key_outputs_table(benchmark_data)
    if_title, if_hdrs, if_rows = build_important_findings_table(benchmark_data)

    print_table(fr_title, fr_hdrs, fr_rows)
    print_table(ko_title, ko_hdrs, ko_rows)
    print_table(if_title, if_hdrs, if_rows)

    # ── Excel output ──
    out_dir = input("\nSave Excel to folder [tables_output]: ").strip()
    out_dir = out_dir if out_dir else "tables_output"
    os.makedirs(out_dir, exist_ok=True)

    fname_parts = "_".join(b[0].replace(" ", "_") for b in benchmark_data)
    out_path = os.path.join(out_dir, f"simscalar_results_{fname_parts}.xlsx")

    summary_tables = [
        (fr_title, fr_hdrs, fr_rows),
        (ko_title, ko_hdrs, ko_rows),
        (if_title, if_hdrs, if_rows),
    ]

    write_excel(out_path, benchmark_data, comp_rows, bnames, summary_tables)

    print(f"\n All done! Open '{out_path}' to view your formatted tables.\n")


if __name__ == "__main__":
    main()
